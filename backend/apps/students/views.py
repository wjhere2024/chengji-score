import csv
from io import BytesIO, TextIOWrapper

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.classes.models import Class
from apps.users.mixins import SchoolScopedQuerysetMixin
from .models import Student
from .serializers import (
    StudentCreateUpdateSerializer,
    StudentDetailSerializer,
    StudentImportSerializer,
    StudentSerializer,
)


class StudentViewSet(SchoolScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = Student.objects.select_related("class_obj", "user").all()
    permission_classes = [IsAuthenticated]
    filterset_fields = {"class_obj": ["exact"], "gender": ["exact"], "is_active": ["exact"]}
    search_fields = ["student_id", "name", "parent_name", "phone"]
    ordering_fields = ["student_id", "name", "created_at"]
    ordering = ["class_obj__grade", "class_obj__class_number", "student_id"]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return StudentDetailSerializer
        if self.action in ["create", "update", "partial_update"]:
            return StudentCreateUpdateSerializer
        if self.action == "import_students":
            return StudentImportSerializer
        return StudentSerializer

    def _allowed_class_queryset(self):
        user = self.request.user
        queryset = Class.objects.all()
        if getattr(user, "school", None):
            queryset = queryset.filter(school=user.school)
        if user.is_admin:
            return queryset
        return queryset.filter(head_teacher=user)

    @action(detail=False, methods=["get"])
    def download_template(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "students"
        ws.append(["学号", "姓名", "性别", "班级", "入学年份", "家长姓名", "联系电话", "备注"])
        ws.append(["2024001", "张三", "男", "一年级1班", 2024, "张先生", "13800138000", ""])
        ws.append(["", "李四", "女", "一年级1班", 2024, "李女士", "13900139000", "学号可留空自动生成"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_import_template.xlsx"'
        return response

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_students(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = serializer.validated_data["file"]
        default_class_id = serializer.validated_data.get("class_id")
        allowed_classes = self._allowed_class_queryset()

        rows = []
        if file.name.endswith(".csv"):
            file.seek(0)
            reader = csv.DictReader(TextIOWrapper(file.file, encoding="utf-8-sig"))
            rows = list(reader)
        else:
            workbook = load_workbook(file, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers))})

        default_class = None
        if default_class_id:
            default_class = allowed_classes.filter(id=default_class_id).first()
            if not default_class:
                return Response({"error": "默认班级无权限或不存在"}, status=status.HTTP_400_BAD_REQUEST)

        created = 0
        updated = 0
        errors = []

        for index, row in enumerate(rows, start=2):
            name = str(row.get("姓名") or row.get("姓名*") or "").strip()
            if not name:
                continue

            class_name = str(row.get("班级") or row.get("班级*") or "").strip()
            class_obj = default_class or allowed_classes.filter(name=class_name).first()
            if not class_obj:
                errors.append(f"第{index}行: 班级不存在或无权限")
                continue

            student_id = str(row.get("学号") or "").strip()
            defaults = {
                "name": name,
                "gender": "female" if str(row.get("性别") or "").strip() in ["女", "female"] else "male",
                "class_obj": class_obj,
                "admission_year": row.get("入学年份") or None,
                "parent_name": str(row.get("家长姓名") or "").strip() or None,
                "phone": str(row.get("联系电话") or "").strip() or None,
                "notes": str(row.get("备注") or "").strip() or None,
            }

            if student_id:
                student, is_created = Student.objects.update_or_create(
                    student_id=student_id,
                    class_obj=class_obj,
                    defaults=defaults,
                )
            else:
                student = Student.objects.create(**defaults)
                is_created = True

            if is_created:
                created += 1
            else:
                updated += 1

        return Response(
            {"message": "导入完成", "created": created, "updated": updated, "errors": errors[:20]}
        )

    @action(detail=False, methods=["get"])
    def export_students(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "students"
        ws.append(["学号", "姓名", "性别", "班级", "入学年份", "家长姓名", "联系电话", "备注"])

        for student in self.filter_queryset(self.get_queryset()):
            ws.append(
                [
                    student.student_id,
                    student.name,
                    student.get_gender_display(),
                    student.class_obj.name,
                    student.admission_year,
                    student.parent_name,
                    student.phone,
                    student.notes,
                ]
            )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
        return response

    @action(detail=False, methods=["get"])
    def my_students(self, request):
        serializer = self.get_serializer(self.get_queryset(), many=True)
        return Response(serializer.data)

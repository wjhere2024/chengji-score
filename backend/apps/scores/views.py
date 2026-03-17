import csv
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO, TextIOWrapper

from django.db.models import Avg, Count, Max, Min, Q
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from pypinyin import Style, lazy_pinyin, pinyin
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.classes.models import Class
from apps.exams.models import Exam, ExamGradeSubject
from apps.students.models import Student
from apps.subjects.models import Subject
from .models import Score
from .serializers import (
    ScoreCreateUpdateSerializer,
    ScoreDetailSerializer,
    ScoreImportSerializer,
    ScoreSerializer,
    ScoreTextImportConfirmSerializer,
    ScoreTextParseSerializer,
)


HEADER_STUDENT_ID = "学号"
HEADER_CANDIDATE_ID = "考号"
HEADER_NAME = "姓名"
HEADER_CLASS = "班级"
HEADER_SCORE = "成绩"


class ScoreViewSet(viewsets.ModelViewSet):
    queryset = Score.objects.select_related("student", "exam", "subject", "created_by", "student__class_obj").all()
    permission_classes = [IsAuthenticated]
    filterset_fields = ["student", "exam", "subject", "student__class_obj"]
    search_fields = ["student__name", "student__student_id"]
    ordering_fields = ["score", "created_at"]
    ordering = ["-exam__exam_date", "student__class_obj__grade", "student__class_obj__class_number"]

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        if getattr(user, "school", None):
            queryset = queryset.filter(student__class_obj__school=user.school)
        if user.is_admin:
            return queryset
        return queryset.filter(student__class_obj__head_teacher=user)

    def get_serializer_class(self):
        if self.action == "retrieve":
            return ScoreDetailSerializer
        if self.action in ["create", "update", "partial_update"]:
            return ScoreCreateUpdateSerializer
        if self.action == "import_scores":
            return ScoreImportSerializer
        if self.action == "parse_text_scores":
            return ScoreTextParseSerializer
        if self.action == "import_text_scores":
            return ScoreTextImportConfirmSerializer
        return ScoreSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=["get"])
    def download_template(self, request):
        exam_id = request.query_params.get("exam_id")
        template_type = request.query_params.get("type", "single")

        workbook = Workbook()
        sheet = workbook.active

        if template_type == "multi":
            headers = [HEADER_STUDENT_ID, HEADER_NAME, HEADER_CLASS]
            if exam_id:
                exam = Exam.objects.filter(id=exam_id).first()
                if exam:
                    names = ExamGradeSubject.objects.filter(exam=exam).select_related("subject")
                    headers.extend(sorted({item.subject.name for item in names}))
            if len(headers) == 3:
                headers.extend(["语文", "数学", "英语"])
            sheet.title = "scores_multi"
            sheet.append(headers)
            sample = ["2024001", "张三", "一年级1班", 95, 96, 97]
            sheet.append(sample[: len(headers)])
        else:
            sheet.title = "scores_single"
            sheet.append([HEADER_STUDENT_ID, HEADER_NAME, HEADER_CLASS, HEADER_SCORE])
            sheet.append(["2024001", "张三", "一年级1班", 95])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="score_import_template.xlsx"'
        return response

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_scores(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = serializer.validated_data["file"]
        exam = Exam.objects.filter(id=serializer.validated_data["exam_id"]).first()
        if not exam:
            return Response({"error": "考试不存在"}, status=status.HTTP_400_BAD_REQUEST)

        subject_id = serializer.validated_data.get("subject_id")
        rows = self._read_rows(file)
        columns = list(rows[0].keys()) if rows else []
        subject_map = self._resolve_subject_columns(columns, subject_id)
        if isinstance(subject_map, Response):
            return subject_map

        result = self._save_rows(request, exam, rows, subject_map)
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_workbook(self, request):
        file = request.FILES.get("file")
        exam_id = request.data.get("exam_id")
        if not file or not exam_id:
            return Response({"error": "file 和 exam_id 必填"}, status=status.HTTP_400_BAD_REQUEST)

        exam = Exam.objects.filter(id=exam_id).first()
        if not exam:
            return Response({"error": "考试不存在"}, status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(file, data_only=True)
        all_classes = list(self._class_queryset(request.user))

        success_count = 0
        errors = []
        processed_sheets = 0

        for sheet_name in workbook.sheetnames:
            class_obj = self._match_class(sheet_name, all_classes)
            if not class_obj:
                continue

            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[3]]
            rows = [{headers[i]: row[i] for i in range(len(headers))} for row in sheet.iter_rows(min_row=4, values_only=True)]
            subject_map = self._resolve_subject_columns(headers, None)
            if isinstance(subject_map, Response):
                continue

            result = self._save_rows(request, exam, rows, subject_map, class_obj=class_obj)
            success_count += result["success_count"]
            errors.extend(result["errors"])
            processed_sheets += 1

        return Response(
            {
                "message": "导入完成",
                "processed_sheets": processed_sheets,
                "success_count": success_count,
                "errors": errors[:50],
            }
        )

    @action(detail=False, methods=["post"])
    def parse_text_scores(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        exam = Exam.objects.filter(id=serializer.validated_data["exam_id"]).first()
        subject = Subject.objects.filter(id=serializer.validated_data["subject_id"]).first()
        if not exam or not subject:
            return Response({"error": "考试或科目不存在"}, status=status.HTTP_400_BAD_REQUEST)

        class_obj = None
        class_id = serializer.validated_data.get("class_id")
        if class_id:
            class_obj = self._class_queryset(request.user).filter(id=class_id).first()
            if not class_obj:
                return Response({"error": "班级不存在或无权限"}, status=status.HTTP_400_BAD_REQUEST)

        records, errors = self._parse_text_records(
            request.user,
            serializer.validated_data["text"],
            subject,
            class_obj=class_obj,
        )
        return Response(
            {
                "matched_count": len(records),
                "error_count": len(errors),
                "records": records,
                "errors": errors,
            }
        )

    @action(detail=False, methods=["post"])
    def import_text_scores(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        exam = Exam.objects.filter(id=serializer.validated_data["exam_id"]).first()
        subject = Subject.objects.filter(id=serializer.validated_data["subject_id"]).first()
        if not exam or not subject:
            return Response({"error": "考试或科目不存在"}, status=status.HTTP_400_BAD_REQUEST)

        students = self._student_queryset(request.user)
        success_count = 0
        errors = []

        for index, record in enumerate(serializer.validated_data["records"], start=1):
            student = students.filter(id=record.get("student_id")).first()
            if not student:
                errors.append(f"第 {index} 条学生不存在或无权限")
                continue

            try:
                score_value = Decimal(str(record.get("score")))
            except (InvalidOperation, TypeError):
                errors.append(f"第 {index} 条分数格式错误")
                continue

            if score_value < 0 or score_value > subject.full_score:
                errors.append(f"第 {index} 条分数超出范围")
                continue

            Score.objects.update_or_create(
                student=student,
                exam=exam,
                subject=subject,
                defaults={
                    "score": score_value,
                    "created_by": request.user,
                    "notes": record.get("notes", "文本批量导入"),
                },
            )
            success_count += 1

        return Response({"message": "导入完成", "success_count": success_count, "errors": errors})

    def _read_rows(self, file):
        if file.name.endswith(".csv"):
            file.seek(0)
            return list(csv.DictReader(TextIOWrapper(file.file, encoding="utf-8-sig")))

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        return [{headers[i]: row[i] for i in range(len(headers))} for row in sheet.iter_rows(min_row=2, values_only=True)]

    def _resolve_subject_columns(self, columns, subject_id):
        if subject_id:
            subject = Subject.objects.filter(id=subject_id).first()
            if not subject:
                return Response({"error": "科目不存在"}, status=status.HTTP_400_BAD_REQUEST)
            return [{"name": subject.name, "subject": subject, "column": HEADER_SCORE}]

        subject_columns = []
        ignored = {HEADER_STUDENT_ID, HEADER_CANDIDATE_ID, HEADER_NAME, HEADER_CLASS, "性别", "序号", None, ""}
        for column in columns:
            if column in ignored:
                continue
            subject = Subject.objects.filter(name=str(column).strip()).first()
            if subject:
                subject_columns.append({"name": subject.name, "subject": subject, "column": column})

        if not subject_columns:
            return Response({"error": "未识别到有效科目列"}, status=status.HTTP_400_BAD_REQUEST)
        return subject_columns

    def _save_rows(self, request, exam, rows, subject_columns, class_obj=None):
        success_count = 0
        errors = []

        for index, row in enumerate(rows, start=2):
            student = self._find_student(request.user, row, class_obj)
            if not student:
                errors.append(f"第 {index} 行未找到学生")
                continue

            for subject_info in subject_columns:
                raw_value = row.get(subject_info["column"])
                if raw_value in [None, "", "None"]:
                    continue

                try:
                    score_value = Decimal(str(raw_value).strip())
                except (InvalidOperation, AttributeError):
                    errors.append(f"第 {index} 行 {subject_info['name']} 分数格式错误")
                    continue

                if score_value < 0 or score_value > subject_info["subject"].full_score:
                    errors.append(f"第 {index} 行 {subject_info['name']} 分数超出范围")
                    continue

                Score.objects.update_or_create(
                    student=student,
                    exam=exam,
                    subject=subject_info["subject"],
                    defaults={"score": score_value, "created_by": request.user},
                )
                success_count += 1

        return {"message": "导入完成", "success_count": success_count, "errors": errors}

    def _student_queryset(self, user):
        queryset = Student.objects.select_related("class_obj").all()
        if getattr(user, "school", None):
            queryset = queryset.filter(class_obj__school=user.school)
        if user.is_admin:
            return queryset
        return queryset.filter(class_obj__head_teacher=user)

    def _class_queryset(self, user):
        queryset = Class.objects.all()
        if getattr(user, "school", None):
            queryset = queryset.filter(school=user.school)
        if user.is_admin:
            return queryset
        return queryset.filter(head_teacher=user)

    def _find_student(self, user, row, matched_class=None):
        queryset = self._student_queryset(user)

        student_id = str(row.get(HEADER_STUDENT_ID) or row.get(HEADER_CANDIDATE_ID) or "").strip()
        if student_id:
            student = queryset.filter(student_id=student_id).first()
            if student:
                return student

        student_name = str(row.get(HEADER_NAME) or "").strip()
        if not student_name:
            return None

        class_name = str(row.get(HEADER_CLASS) or "").strip()
        class_obj = matched_class or self._match_class(class_name, list(self._class_queryset(user)))
        if class_obj:
            return queryset.filter(name=student_name, class_obj=class_obj).first()

        students = queryset.filter(name=student_name)
        if students.count() == 1:
            return students.first()
        return None

    def _parse_text_records(self, user, text, subject, class_obj=None):
        records = []
        errors = []

        for index, raw_line in enumerate(self._split_text_lines(text), start=1):
            parsed = self._extract_name_score(raw_line)
            if not parsed:
                errors.append(f"第 {index} 行无法识别: {raw_line}")
                continue

            student = self._match_student_by_name(user, parsed["name"], class_obj=class_obj)
            if not student:
                errors.append(f"第 {index} 行未匹配到学生: {raw_line}")
                continue

            if parsed["score"] < 0 or parsed["score"] > subject.full_score:
                errors.append(f"第 {index} 行分数超出范围: {raw_line}")
                continue

            records.append(
                {
                    "line_no": index,
                    "raw_line": raw_line,
                    "student_id": student.id,
                    "student_name": student.name,
                    "class_id": student.class_obj_id,
                    "class_name": student.class_obj.name,
                    "score": f"{parsed['score']:.1f}",
                    "matched_by": parsed["matched_by"],
                }
            )

        return records, errors

    def _split_text_lines(self, text):
        normalized = str(text or "").replace("\r", "\n")
        chunks = []
        for block in normalized.split("\n"):
            block = block.strip()
            if not block:
                continue
            for piece in re.split(r"[；;。]+", block):
                piece = piece.strip(" ,，、")
                if piece:
                    chunks.append(piece)
        return chunks

    def _extract_name_score(self, raw_line):
        line = str(raw_line).strip()
        patterns = [
            r"^(?P<name>[\u4e00-\u9fa5A-Za-z0-9·]{1,20})[\s,:：，、-]+(?P<score>[0-9]{1,3}(?:\.[0-9])?|[零一二三四五六七八九十百两点〇○]+)(?:分)?$",
            r"^(?P<name>[\u4e00-\u9fa5A-Za-z0-9·]{1,20})(?P<score>[0-9]{1,3}(?:\.[0-9])?)(?:分)?$",
        ]
        for pattern in patterns:
            matched = re.match(pattern, line)
            if not matched:
                continue
            score = self._parse_score_value(matched.group("score"))
            if score is None:
                return None
            return {"name": matched.group("name").strip(), "score": score, "matched_by": "name_or_pinyin"}
        return None

    def _parse_score_value(self, raw_value):
        value = str(raw_value).strip()
        try:
            return Decimal(value)
        except InvalidOperation:
            pass

        digits = {"零": 0, "〇": 0, "○": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}
        if "点" in value:
            integer_part, decimal_part = value.split("点", 1)
            integer_value = self._parse_chinese_integer(integer_part, digits)
            if integer_value is None:
                return None
            decimal_digits = []
            for char in decimal_part:
                if char not in digits:
                    return None
                decimal_digits.append(str(digits[char]))
            return Decimal(f"{integer_value}.{''.join(decimal_digits)}")

        integer_value = self._parse_chinese_integer(value, digits)
        if integer_value is None:
            return None
        return Decimal(integer_value)

    def _parse_chinese_integer(self, value, digits):
        if not value:
            return 0
        if value == "十":
            return 10

        total = 0
        current = 0
        units = {"十": 10, "百": 100}
        for char in value:
            if char in digits:
                current = digits[char]
            elif char in units:
                if current == 0:
                    current = 1
                total += current * units[char]
                current = 0
            else:
                return None
        return total + current

    def _match_student_by_name(self, user, raw_name, class_obj=None):
        target = str(raw_name or "").strip()
        if not target:
            return None

        queryset = self._student_queryset(user)
        if class_obj:
            queryset = queryset.filter(class_obj=class_obj)

        exact = queryset.filter(name=target)
        if exact.count() == 1:
            return exact.first()

        target_lower = target.lower()
        exact_pinyin = []
        fuzzy_pinyin = []
        for student in queryset:
            full_pinyin = "".join(lazy_pinyin(student.name)).lower()
            initials = "".join(item[0][0] for item in pinyin(student.name, style=Style.FIRST_LETTER) if item and item[0]).lower()
            if target_lower in {full_pinyin, initials}:
                exact_pinyin.append(student)
            elif full_pinyin.startswith(target_lower) or initials.startswith(target_lower):
                fuzzy_pinyin.append(student)

        if len(exact_pinyin) == 1:
            return exact_pinyin[0]
        if not exact.exists() and len(fuzzy_pinyin) == 1:
            return fuzzy_pinyin[0]
        return None

    def _match_class(self, class_name, all_classes):
        class_name = str(class_name or "").strip()
        if not class_name:
            return None
        for class_obj in all_classes:
            if class_obj.name == class_name or class_name in class_obj.name or class_obj.name.startswith(class_name):
                return class_obj
        return None

    @action(detail=False, methods=["get"])
    def statistics(self, request):
        exam_id = request.query_params.get("exam_id")
        subject_id = request.query_params.get("subject_id")
        class_id = request.query_params.get("class_id")
        queryset = self.get_queryset()
        if exam_id:
            queryset = queryset.filter(exam_id=exam_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if class_id:
            queryset = queryset.filter(student__class_obj_id=class_id)

        stats = queryset.aggregate(
            total_count=Count("id"),
            average_score=Avg("score"),
            max_score=Max("score"),
            min_score=Min("score"),
            pass_count=Count("id", filter=Q(score__gte=60)),
            excellent_count=Count("id", filter=Q(score__gte=90)),
        )
        total_count = stats["total_count"] or 0
        pass_count = stats["pass_count"] or 0
        excellent_count = stats["excellent_count"] or 0

        return Response(
            {
                "total_count": total_count,
                "average_score": stats["average_score"] or 0,
                "max_score": stats["max_score"] or 0,
                "min_score": stats["min_score"] or 0,
                "pass_count": pass_count,
                "pass_rate": round(pass_count * 100 / total_count, 2) if total_count else 0,
                "excellent_count": excellent_count,
                "excellent_rate": round(excellent_count * 100 / total_count, 2) if total_count else 0,
            }
        )

    @action(detail=False, methods=["get"])
    def student_history(self, request):
        student_id = request.query_params.get("student_id")
        if not student_id:
            return Response({"error": "student_id 必填"}, status=status.HTTP_400_BAD_REQUEST)
        queryset = self.get_queryset().filter(student_id=student_id).order_by("-exam__exam_date")
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
